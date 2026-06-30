import datetime
import logging
import random
import re

import openpyxl
from django.db.models import Count
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.courses.models import Level, Lesson
from apps.exercises.constants import ANTICHEAT_ENFORCE_KINDS, MAX_ATTEMPTS_PER_QUESTION, MAX_SESSION_SECONDS, MIN_ANSWER_MS, SKIP_ANSWER_SENTINEL
from apps.exercises.generators.curated import CuratedGenerator
from apps.exercises.generators.procedural import ProceduralGenerator
from apps.progress.models import LessonCompletion, LevelCompletion, ProgressRecord
from apps.progress.services import finalize_session, record_attempt
from apps.users.permissions import IsAdmin

from .models import ArenaSession, CuratedQuestion, ExerciseTemplate, SessionKind
from .serializers import AttemptSerializer, ProgressRecordSerializer, SessionMetaSerializer

_log = logging.getLogger("apps.exercises.anticheat")


class StartClassworkView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, level_id):
        level = get_object_or_404(Level, pk=level_id)

        if level.order > 1:
            prev_level = Level.objects.filter(order=level.order - 1).first()
            if prev_level and not LevelCompletion.objects.filter(
                user=request.user, level=prev_level, kind="CLASSWORK"
            ).exists():
                return Response({"detail": "Level is locked."}, status=status.HTTP_403_FORBIDDEN)

        existing = ArenaSession.objects.filter(
            user=request.user,
            kind=SessionKind.CLASSWORK,
            template__lesson__level=level,
            submitted_at__isnull=True,
            abandoned_at__isnull=True,
        ).first()
        if existing:
            return Response(SessionMetaSerializer(existing).data)

        template = (
            ExerciseTemplate.objects.filter(lesson__level=level, kind=SessionKind.CLASSWORK)
            .select_related("lesson")
            .first()
        )
        if template is None:
            return Response(
                {"detail": "No content available for this level."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        seed = random.getrandbits(63)
        generator = CuratedGenerator(seed=seed, template_config=template.config_json)
        questions = generator.generate()
        is_test_mode = bool(request.data.get("is_test_mode", False))

        session = ArenaSession.objects.create(
            user=request.user,
            kind=SessionKind.CLASSWORK,
            template=template,
            config_json=template.config_json,
            seed=seed,
            questions_json=[q.to_dict() for q in questions],
            is_test_mode=is_test_mode,
        )
        return Response(SessionMetaSerializer(session).data, status=status.HTTP_201_CREATED)


class StartLessonClassworkView(APIView):
    """
    POST /levels/{level_id}/lessons/{lesson_id}/classwork/start/
    Lesson-scoped classwork start. Enforces lesson-level locking via LessonCompletion.
    Returns 200 on resume, 201 on new session, 403 if locked, 400 if no template.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, level_id, lesson_id):
        lesson = get_object_or_404(Lesson, pk=lesson_id, level__id=level_id)

        if lesson.order > 1:
            prev_lesson = Lesson.objects.filter(
                level=lesson.level, order=lesson.order - 1
            ).first()
            if prev_lesson and not LessonCompletion.objects.filter(
                user=request.user, lesson=prev_lesson, kind="CLASSWORK"
            ).exists():
                return Response({"detail": "Lesson is locked."}, status=status.HTTP_403_FORBIDDEN)

        existing = ArenaSession.objects.filter(
            user=request.user,
            kind=SessionKind.CLASSWORK,
            template__lesson=lesson,
            submitted_at__isnull=True,
            abandoned_at__isnull=True,
        ).first()
        if existing:
            return Response(SessionMetaSerializer(existing).data)

        template = ExerciseTemplate.objects.filter(
            lesson=lesson, kind=SessionKind.CLASSWORK
        ).first()
        if template is None:
            return Response(
                {"detail": "No content available for this lesson."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        seed = random.getrandbits(63)
        generator = CuratedGenerator(seed=seed, template_config=template.config_json)
        questions = generator.generate()
        is_test_mode = bool(request.data.get("is_test_mode", False))

        session = ArenaSession.objects.create(
            user=request.user,
            kind=SessionKind.CLASSWORK,
            template=template,
            config_json=template.config_json,
            seed=seed,
            questions_json=[q.to_dict() for q in questions],
            is_test_mode=is_test_mode,
        )
        return Response(SessionMetaSerializer(session).data, status=status.HTTP_201_CREATED)


class StartPracticeView(APIView):
    permission_classes = [IsAuthenticated]

    VALID_MODES = {"TIME_ATTACK", "ZEN", "CUSTOM", "FLASH_CARDS"}
    # MIXED is practice-only: the generator resolves it to a per-question ADD/SUB mix.
    VALID_OPERATIONS = {"ADD", "SUB", "MUL", "DIV", "MIXED"}

    def post(self, request):
        errors = {}

        mode = request.data.get("mode")
        if not mode:
            errors["mode"] = "Required."
        elif mode not in self.VALID_MODES:
            errors["mode"] = f"Must be one of {sorted(self.VALID_MODES)}."

        operation = request.data.get("operation")
        if not operation:
            errors["operation"] = "Required."
        elif operation not in self.VALID_OPERATIONS:
            errors["operation"] = f"Must be one of {sorted(self.VALID_OPERATIONS)}."
        elif mode == "FLASH_CARDS" and operation in {"MUL", "DIV"}:
            # D-5 constraint: flash cards are add/sub drills only; MUL/DIV is unsupported.
            errors["operation"] = "Flash Cards mode only supports ADD, SUB, or MIXED."

        flash_speed_ms = None
        if mode == "FLASH_CARDS":
            try:
                flash_speed_ms = int(request.data.get("flash_speed_ms", 2000))
                if not 500 <= flash_speed_ms <= 10000:
                    errors["flash_speed_ms"] = "Must be between 500 and 10000."
            except (TypeError, ValueError):
                errors["flash_speed_ms"] = "Must be an integer."

        try:
            digits = int(request.data.get("digits", ""))
            if not 1 <= digits <= 4:
                errors["digits"] = "Must be between 1 and 4."
        except (TypeError, ValueError):
            errors["digits"] = "Must be an integer."

        # Per-row digit config for MUL/DIV (D-6). Only validated + stored when present.
        digits_row1 = None
        digits_row2 = None
        if request.data.get("digits_row1") is not None:
            try:
                digits_row1 = int(request.data["digits_row1"])
                if not 1 <= digits_row1 <= 4:
                    errors["digits_row1"] = "Must be between 1 and 4."
            except (TypeError, ValueError):
                errors["digits_row1"] = "Must be an integer."
        if request.data.get("digits_row2") is not None:
            try:
                digits_row2 = int(request.data["digits_row2"])
                if not 1 <= digits_row2 <= 2:
                    errors["digits_row2"] = "Must be between 1 and 2."
            except (TypeError, ValueError):
                errors["digits_row2"] = "Must be an integer."

        try:
            rows = int(request.data.get("rows", ""))
            if not 2 <= rows <= 8:
                errors["rows"] = "Must be between 2 and 8."
        except (TypeError, ValueError):
            errors["rows"] = "Must be an integer."

        try:
            question_count = int(request.data.get("question_count", ""))
            if not 1 <= question_count <= 100:
                errors["question_count"] = "Must be between 1 and 100."
        except (TypeError, ValueError):
            errors["question_count"] = "Must be an integer."

        try:
            time_limit_sec = int(request.data.get("time_limit_sec", ""))
            if not (0 <= time_limit_sec <= MAX_SESSION_SECONDS):
                errors["time_limit_sec"] = f"Must be between 0 and {MAX_SESSION_SECONDS}."
        except (TypeError, ValueError):
            errors["time_limit_sec"] = "Must be an integer."

        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        config = {
            "operation": operation,
            "digits": digits,
            "rows": rows,
            "question_count": question_count,
            "time_limit_sec": time_limit_sec,
        }
        if flash_speed_ms is not None:
            config["flash_speed_ms"] = flash_speed_ms
        if operation in {"MUL", "DIV"}:
            if digits_row1 is not None:
                config["digits_row1"] = digits_row1
            if digits_row2 is not None:
                config["digits_row2"] = digits_row2

        seed = random.getrandbits(63)
        questions = ProceduralGenerator(seed=seed, config=config).generate()

        session = ArenaSession.objects.create(
            user=request.user,
            kind=mode,
            template=None,
            config_json=config,
            seed=seed,
            questions_json=[q.to_dict() for q in questions],
        )
        return Response(SessionMetaSerializer(session).data, status=status.HTTP_201_CREATED)


class SessionDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, session_id):
        session = get_object_or_404(ArenaSession, pk=session_id, user=request.user)
        return Response(SessionMetaSerializer(session).data)


class SubmitAttemptView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        # select_for_update serializes concurrent submissions for the same session,
        # making count()+1 for attempt_number atomic within ATOMIC_REQUESTS transaction.
        session = get_object_or_404(
            ArenaSession.objects.select_for_update(),
            pk=session_id,
            user=request.user,
        )

        if not session.is_active:
            return Response({"detail": "Session is not active."}, status=status.HTTP_400_BAD_REQUEST)

        question_index = request.data.get("question_index")
        if question_index is None:
            return Response({"detail": "question_index is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            question_index = int(question_index)
        except (TypeError, ValueError):
            return Response({"detail": "question_index must be an integer."}, status=status.HTTP_400_BAD_REQUEST)

        if question_index < 0 or question_index >= len(session.questions_json):
            return Response({"detail": "question_index out of bounds."}, status=status.HTTP_400_BAD_REQUEST)

        raw_answer = request.data.get("answer")
        if raw_answer is None:
            return Response({"detail": "answer is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            answer = int(raw_answer)
        except (TypeError, ValueError):
            return Response({"detail": "answer must be an integer."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            elapsed_ms = int(request.data.get("elapsed_ms", 0))
        except (TypeError, ValueError):
            elapsed_ms = 0
        if elapsed_ms < 0:
            elapsed_ms = 0

        is_skip = answer == SKIP_ANSWER_SENTINEL

        if not is_skip and elapsed_ms < MIN_ANSWER_MS:
            _log.warning(
                "min_answer_ms_violation",
                extra={
                    "session_id": str(session.id),
                    "user_id": str(request.user.id),
                    "question_index": question_index,
                    "elapsed_ms": elapsed_ms,
                    "session_kind": session.kind,
                },
            )
            if session.kind in ANTICHEAT_ENFORCE_KINDS:
                return Response(
                    {"detail": "Submission rejected: answer submitted implausibly fast."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        from apps.progress.models import QuestionAttempt

        attempt_number = (
            QuestionAttempt.objects.filter(
                session=session, question_index=question_index
            ).count()
            + 1
        )

        if session.is_test_mode:
            if is_skip:
                return Response(
                    {"detail": "Skipping is not allowed in Test Mode."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if attempt_number > 1:
                return Response(
                    {"detail": "Only one attempt per question in Test Mode."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        q = session.questions_json[question_index]
        attempt = record_attempt(
            session=session,
            question_index=question_index,
            attempt_number=attempt_number,
            question_text=q["text"],
            expected_answer=q["answer"],
            submitted_answer=answer,
            elapsed_ms=elapsed_ms,
            is_skip=is_skip,
        )

        return Response({
            "question_index": attempt.question_index,
            "is_correct": attempt.is_correct,
            "xp_delta": 0,
        })


class BulkSubmitAttemptView(APIView):
    """
    POST /sessions/{id}/attempts/bulk/
    Submit multiple attempts in one call.  Intended for:
      - Practice: client-side grading sessions submitting all answers at once.
      - Classwork: debounced background flushes to protect against data loss.

    Idempotent on (session, question_index, attempt_number): if an attempt for the
    computed attempt_number already exists, its verdict is returned without a new row.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        session = get_object_or_404(
            ArenaSession.objects.select_for_update(),
            pk=session_id,
            user=request.user,
        )

        if not session.is_active:
            return Response({"detail": "Session is not active."}, status=status.HTTP_400_BAD_REQUEST)

        attempts_data = request.data.get("attempts")
        if not isinstance(attempts_data, list):
            return Response({"detail": "attempts must be a list."}, status=status.HTTP_400_BAD_REQUEST)

        if not attempts_data:
            return Response({"verdicts": []})

        q_count = len(session.questions_json)
        seen_indices: set[int] = set()
        for i, item in enumerate(attempts_data):
            try:
                idx = int(item.get("question_index"))
            except (TypeError, ValueError):
                return Response(
                    {"detail": f"Item {i}: question_index must be an integer."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if idx < 0 or idx >= q_count:
                return Response(
                    {"detail": f"Item {i}: question_index {idx} out of bounds (0..{q_count - 1})."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if idx in seen_indices:
                return Response(
                    {"detail": f"Item {i}: duplicate question_index {idx}. Each index may appear once per bulk call."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            seen_indices.add(idx)

        from apps.progress.models import QuestionAttempt

        # Reject if any question has already reached the attempt cap. Checked once
        # before the loop so the entire batch is rejected rather than partially applied.
        existing_counts = {
            row["question_index"]: row["cnt"]
            for row in QuestionAttempt.objects.filter(
                session=session,
                question_index__in=seen_indices,
            ).values("question_index").annotate(cnt=Count("id"))
        }
        for idx in seen_indices:
            if existing_counts.get(idx, 0) >= MAX_ATTEMPTS_PER_QUESTION:
                return Response(
                    {"detail": f"question_index {idx} has reached the maximum of {MAX_ATTEMPTS_PER_QUESTION} attempts."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        verdicts = []
        for item in attempts_data:
            question_index = int(item["question_index"])
            try:
                answer = int(item["answer"])
            except (TypeError, ValueError):
                return Response({"detail": "answer must be an integer."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                elapsed_ms = int(item.get("elapsed_ms", 0))
            except (TypeError, ValueError):
                elapsed_ms = 0
            elapsed_ms = max(0, elapsed_ms)

            is_skip = answer == SKIP_ANSWER_SENTINEL

            if not is_skip and elapsed_ms < MIN_ANSWER_MS:
                _log.warning(
                    "min_answer_ms_violation",
                    extra={
                        "session_id": str(session.id),
                        "user_id": str(request.user.id),
                        "question_index": question_index,
                        "elapsed_ms": elapsed_ms,
                        "session_kind": session.kind,
                        "endpoint": "bulk",
                    },
                )
                if session.kind in ANTICHEAT_ENFORCE_KINDS:
                    # Skip this attempt rather than abort the whole bulk call;
                    # the other attempts in the batch are still recorded.
                    verdicts.append({
                        "question_index": question_index,
                        "is_correct": False,
                        "xp_delta": 0,
                        "rejected": True,
                    })
                    continue

            attempt_number = (
                QuestionAttempt.objects.filter(
                    session=session, question_index=question_index
                ).count()
                + 1
            )

            q = session.questions_json[question_index]
            existing = QuestionAttempt.objects.filter(
                session=session,
                question_index=question_index,
                attempt_number=attempt_number,
            ).first()

            if existing is not None:
                attempt = existing
            else:
                attempt = record_attempt(
                    session=session,
                    question_index=question_index,
                    attempt_number=attempt_number,
                    question_text=q["text"],
                    expected_answer=q["answer"],
                    submitted_answer=answer,
                    elapsed_ms=elapsed_ms,
                    is_skip=is_skip,
                )

            verdicts.append({
                "question_index": question_index,
                "is_correct": attempt.is_correct,
                "xp_delta": 0,
            })

        return Response({"verdicts": verdicts})


class FinalizeSessionView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        session = get_object_or_404(
            ArenaSession.objects.select_related("template__lesson__level"),
            pk=session_id,
            user=request.user,
        )

        try:
            record = session.progress_record
            return Response(ProgressRecordSerializer(record).data)
        except ProgressRecord.DoesNotExist:
            pass

        try:
            record = finalize_session(session)
        except ValueError as e:
            if "already finalized" not in str(e):
                raise  # unexpected — surface with traceback rather than swallow
            # Race: another request finalized between our check and our call.
            session.refresh_from_db()
            try:
                record = session.progress_record
            except ProgressRecord.DoesNotExist:
                return Response({"detail": "Session finalization failed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(ProgressRecordSerializer(record).data)


class SessionReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, session_id):
        session = get_object_or_404(
            ArenaSession.objects.select_related("template__lesson"),
            pk=session_id,
            user=request.user,
        )

        try:
            record = session.progress_record
        except ProgressRecord.DoesNotExist:
            return Response(
                {"detail": "Session not yet finalized."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        attempts = list(session.attempts.order_by("question_index", "attempt_number"))
        lesson_id = str(session.template.lesson.id) if session.template else None

        # Group attempts by question_index to derive per-question verdict.
        groups: dict[int, list] = {}
        for a in attempts:
            groups.setdefault(a.question_index, []).append(a)

        question_verdicts: dict[int, str] = {}
        for q_idx, group in groups.items():
            if group[0].is_skip:
                question_verdicts[q_idx] = "skipped"
            elif group[-1].is_correct:
                question_verdicts[q_idx] = "fixed" if not group[0].is_correct else "correct"
            else:
                question_verdicts[q_idx] = "wrong"

        return Response({
            "progress": ProgressRecordSerializer(record).data,
            "attempts": AttemptSerializer(attempts, many=True).data,
            "lesson_id": lesson_id,
            "question_verdicts": question_verdicts,
        })


_SHEET_NAME_RE = re.compile(r"^L(\d+)\s+C(\d+)$", re.IGNORECASE)
_import_log = logging.getLogger("apps.exercises.import")


class ImportQuestionsView(APIView):
    """
    ADMIN-only: import questions from BOLT ALL LEVELS DATASET (.xlsx).

    Sheet naming: "L{n} C{m}" → Level n, Class m.
    Row schema: row_index | topic_name | question | operator | answer
    Corrupted rows (Excel date serials in question column) are skipped with a warning.
    Each import replaces all existing CuratedQuestion records for the affected lessons (idempotent).
    """

    permission_classes = [IsAuthenticated, IsAdmin]

    def post(self, request):
        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        except Exception as exc:
            return Response(
                {"detail": f"Could not open file: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = []
        for sheet_name in wb.sheetnames:
            m = _SHEET_NAME_RE.match(sheet_name.strip())
            if not m:
                continue  # Skip "Curriculum" and other non-matching sheets

            level_order = int(m.group(1))
            lesson_order = int(m.group(2))
            ws = wb[sheet_name]

            level, _ = Level.objects.get_or_create(
                order=level_order,
                defaults={"name": f"Level {level_order}"},
            )
            lesson, _ = Lesson.objects.get_or_create(
                level=level,
                order=lesson_order,
                defaults={"name": f"Class {lesson_order}"},
            )

            # Replace existing: delete all current curated questions for this lesson.
            deleted_count, _ = CuratedQuestion.objects.filter(lesson=lesson).delete()

            questions_to_create = []
            skipped = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                row_index, topic_name, question, operator, answer = (list(row) + [None] * 5)[:5]

                # Detect Excel date serials: openpyxl returns datetime objects for date-formatted cells.
                if isinstance(question, (datetime.date, datetime.datetime)):
                    msg = f"{sheet_name} row {row_index}: question column is a date serial — skipped"
                    skipped.append(msg)
                    _import_log.warning(msg)
                    continue

                if question is None:
                    continue

                try:
                    answer_int = int(float(answer)) if answer is not None else 0
                except (TypeError, ValueError):
                    msg = f"{sheet_name} row {row_index}: unparseable answer '{answer}' — skipped"
                    skipped.append(msg)
                    _import_log.warning(msg)
                    continue

                questions_to_create.append(CuratedQuestion(
                    lesson=lesson,
                    row_index=int(row_index) if row_index is not None else 0,
                    topic_name=str(topic_name or "")[:256],
                    question_text=str(question)[:512],
                    operator=str(operator or "")[:16],
                    answer=answer_int,
                ))

            CuratedQuestion.objects.bulk_create(questions_to_create, ignore_conflicts=True)
            results.append({
                "sheet": sheet_name,
                "level": level_order,
                "class": lesson_order,
                "imported": len(questions_to_create),
                "replaced": deleted_count,
                "skipped": skipped,
            })

        return Response({"sheets": results, "total_sheets": len(results)})
